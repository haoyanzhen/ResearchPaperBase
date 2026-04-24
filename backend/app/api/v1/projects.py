import io
import logging
import os
import zipfile
from datetime import date, datetime, timezone

import httpx
from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.paper import Paper, ProjectPaperRelation
from app.models.user import User
from app.schemas.common import ok
from app.schemas.project import (
    AddPaperRequest,
    AddPaperResponse,
    CreateProjectRequest,
    CreateRecommendationRequest,
    ExportRequest,
    ExportResponse,
    ModeSwitchRequest,
    ModeSwitchResponse,
    PaperDetail,
    PaperItem,
    PaperListResponse,
    PaperScoreResponse,
    ProjectDetail,
    ProjectListResponse,
    ProjectSummary,
    RecommendationItem,
    RecommendationListResponse,
    ScheduleConfigRequest,
    ScheduleConfigResponse,
    UpdatePaperScoreRequest,
    UpdateProjectRequest,
)
from app.services import project_service
from app.utils.ids import new_id

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/projects", tags=["研究主题"])

_VALID_MODES = {"construction", "deep_research", "review"}


def _project_to_summary(p) -> ProjectSummary:
    return ProjectSummary(
        project_id=p.id,
        name=p.name,
        mode=p.mode,
        status=p.status,
        total_papers=p.total_papers,
        valid_papers=p.valid_papers,
        created_at=p.created_at,
        updated_at=p.updated_at,
    )


def _project_to_detail(p) -> ProjectDetail:
    return ProjectDetail(
        project_id=p.id,
        name=p.name,
        description=p.description,
        mode=p.mode,
        status=p.status,
        total_papers=p.total_papers,
        valid_papers=p.valid_papers,
        auto_push=p.auto_push,
        push_interval=p.push_interval,
        last_push_at=p.last_push_at,
        next_push_at=p.next_push_at,
        created_at=p.created_at,
        updated_at=p.updated_at,
    )


# ── Project CRUD ───────────────────────────────────────────────────────────────

@router.get("")
async def list_projects(
    status: str | None = Query(None),
    mode: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    total, items = await project_service.list_projects(db, current_user.id, status, mode, page, page_size)
    return ok(ProjectListResponse(
        total=total, page=page, page_size=page_size,
        items=[_project_to_summary(p) for p in items],
    ))


@router.post("", status_code=status.HTTP_201_CREATED)
async def create_project(
    body: CreateProjectRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    project = await project_service.create_project(db, current_user.id, body.name, body.description)
    return ok(_project_to_detail(project))


@router.get("/{project_id}")
async def get_project(
    project_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")
    return ok(_project_to_detail(project))


@router.patch("/{project_id}")
async def update_project(
    project_id: str,
    body: UpdateProjectRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")
    project = await project_service.update_project(db, project, body.name, body.description)
    return ok(_project_to_detail(project))


@router.delete("/{project_id}")
async def delete_project(
    project_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")
    await project_service.delete_project(db, project)
    return ok({"message": "deleted"})


# ── FR-005 模式切换 ────────────────────────────────────────────────────────────

@router.post("/{project_id}/mode")
async def switch_mode(
    project_id: str,
    body: ModeSwitchRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if body.target_mode not in _VALID_MODES:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "无效的目标模式")

    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")
    if project.status == "archived":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "已归档课题不允许操作")

    # 运行中任务拦截
    if project.status == "running":
        if body.running_task_action == "leave_and_cancel":
            # 将该课题最新的 running/paused stage_record 标记为 failed
            from sqlalchemy import select as _select
            from app.models.stage import StageRecord as _StageRecord
            from datetime import datetime as _dt, timezone as _tz
            active_rec = await db.execute(
                _select(_StageRecord)
                .where(
                    _StageRecord.project_id == project_id,
                    _StageRecord.status.in_(["running", "paused"]),
                )
                .order_by(_StageRecord.started_at.desc())
                .limit(1)
            )
            active = active_rec.scalar_one_or_none()
            if active:
                active.status = "failed"
                active.error = "用户切换模式时取消"
                active.failed_at = _dt.now(_tz.utc)
            project.status = "idle"
        else:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                "当前有任务正在运行，请选择留在页面或退出任务后切换",
            )

    # 论文库完整性拦截
    if body.target_mode in ("deep_research", "review") and project.valid_papers == 0:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "论文库尚未构建（有效论文数为0），请先完成构建模式",
        )

    previous_mode = project.mode
    project = await project_service.switch_mode(db, project, body.target_mode)
    return ok(ModeSwitchResponse(
        project_id=project.id,
        previous_mode=previous_mode,
        current_mode=project.mode,
        switched_at=datetime.now(timezone.utc),
    ))


# ── FR-006 检索历史查看 ────────────────────────────────────────────────────────

@router.get("/{project_id}/stage-records")
async def get_stage_records(
    project_id: str,
    mode: str | None = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select
    from app.models.stage import StageRecord

    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")

    query = select(StageRecord).where(StageRecord.project_id == project_id)
    if mode:
        query = query.where(StageRecord.mode == mode)
    query = query.order_by(StageRecord.started_at.desc())
    result = await db.execute(query)
    records = result.scalars().all()
    return ok([
        {
            "record_id": r.id,
            "mode": r.mode,
            "stage": r.stage,
            "status": r.status,
            "started_at": r.started_at,
            "completed_at": r.completed_at,
            "result": r.result,
            "error": r.error,
        }
        for r in records
    ])


# ── FR-007 论文库查看与管理 ────────────────────────────────────────────────────

@router.get("/{project_id}/papers")
async def list_papers(
    project_id: str,
    is_valid: bool | None = Query(None),
    push_status: bool | None = Query(None),
    keyword: str | None = Query(None),
    order_by: str = Query("created_at"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")

    total, relations = await project_service.list_papers(
        db, project_id, is_valid, push_status, keyword, order_by, page, page_size
    )
    # 加载关联 paper 数据
    from sqlalchemy import select
    from app.models.paper import Paper

    paper_ids = [r.paper_id for r in relations]
    papers_result = await db.execute(select(Paper).where(Paper.id.in_(paper_ids)))
    papers_map = {p.id: p for p in papers_result.scalars().all()}

    items = []
    for rel in relations:
        p = papers_map.get(rel.paper_id)
        if not p:
            continue
        items.append(PaperItem(
            paper_id=p.id,
            title=p.title,
            authors=p.authors,
            pub_date=str(p.pub_date) if p.pub_date else None,
            venue=p.venue,
            abstract=p.abstract,
            total_score=rel.total_score,
            reference_score=rel.reference_score,
            technical_score=rel.technical_score,
            is_valid=rel.is_valid,
            push_status=rel.push_status,
            download_status=p.download_status,
            ai_analysis=p.ai_analysis,
        ))

    return ok(PaperListResponse(total=total, items=items))


@router.post("/{project_id}/papers", status_code=status.HTTP_201_CREATED)
async def add_paper(
    project_id: str,
    body: AddPaperRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """手动添加论文（DOI / arXiv ID / PDF base64，三选一）—— FR-007"""
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")

    if not any([body.doi, body.arxiv_id, body.pdf_file]):
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "须提供 doi、arxiv_id 或 pdf_file 之一")

    paper_data: dict | None = None

    # ── arXiv ID 获取元数据 ────────────────────────────────────────────────────
    if body.arxiv_id:
        arxiv_id = body.arxiv_id.strip()
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                resp = await client.get(
                    f"https://export.arxiv.org/api/query?id_list={arxiv_id}"
                )
            if resp.status_code == 200:
                from app.agents.construction.stage2_retrieval import _parse_arxiv
                parsed = _parse_arxiv(resp.text, "")
                if parsed:
                    paper_data = {**parsed[0], "source": "arxiv"}
        except Exception as exc:
            logger.warning("arXiv fetch failed for %s: %s", arxiv_id, exc)
        if paper_data is None:
            raise HTTPException(
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                f"无法从 arXiv 获取 {arxiv_id} 的元数据，请检查 ID 是否正确",
            )

    # ── DOI → OpenAlex 获取元数据 ─────────────────────────────────────────────
    elif body.doi:
        doi = body.doi.strip()
        try:
            async with httpx.AsyncClient(timeout=15.0) as client:
                resp = await client.get(
                    f"https://api.openalex.org/works/https://doi.org/{doi}",
                    headers={"User-Agent": "PaperPush/1.0 (mailto:app@paperpush.local)"},
                )
            if resp.status_code == 200:
                w = resp.json()
                title = (w.get("title") or "").strip()
                if title:
                    authors = [
                        (a.get("author", {}) or {}).get("display_name", "")
                        for a in (w.get("authorships") or [])
                    ]
                    pub_year = w.get("publication_year")
                    venue_obj = w.get("primary_location", {}) or {}
                    venue_source = (venue_obj.get("source") or {}) or {}
                    paper_data = {
                        "title": title,
                        "doi": doi,
                        "arxiv_id": None,
                        "authors": [a for a in authors if a],
                        "pub_date": f"{pub_year}-01-01" if pub_year else None,
                        "venue": venue_source.get("display_name") or "OpenAlex",
                        "abstract": w.get("abstract") or None,
                        "source": "openalex",
                    }
        except Exception as exc:
            logger.warning("OpenAlex fetch failed for doi %s: %s", doi, exc)
        if paper_data is None:
            raise HTTPException(
                status.HTTP_422_UNPROCESSABLE_ENTITY,
                f"无法通过 DOI {doi} 获取论文元数据",
            )

    # ── PDF base64 — 保存文件，元数据待后续解析 ────────────────────────────────
    elif body.pdf_file:
        import base64
        paper_id_tmp = new_id("paper")
        try:
            pdf_bytes = base64.b64decode(body.pdf_file)
        except Exception:
            raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "pdf_file 不是有效的 base64 编码")
        pdf_dir = os.path.join(settings.STORAGE_DIR, "papers")
        os.makedirs(pdf_dir, exist_ok=True)
        pdf_path = os.path.join(pdf_dir, f"{paper_id_tmp}.pdf")
        with open(pdf_path, "wb") as fh:
            fh.write(pdf_bytes)
        paper_data = {
            "title": f"（待解析）{paper_id_tmp}",
            "doi": None,
            "arxiv_id": None,
            "authors": [],
            "pub_date": None,
            "venue": None,
            "abstract": None,
            "source": "manual",
            "_pdf_path": pdf_path,
        }

    # ── 写入数据库（ON CONFLICT DO NOTHING 去重）────────────────────────────────
    now = datetime.now(timezone.utc)
    title_lower = (paper_data.get("title") or "").strip().lower()
    if not title_lower:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "无法获取论文标题")

    pub_date_obj: date | None = None
    if paper_data.get("pub_date"):
        try:
            pub_date_obj = date.fromisoformat(paper_data["pub_date"])
        except ValueError:
            pass

    paper_id = new_id("paper")
    insert_stmt = (
        pg_insert(Paper)
        .values(
            id=paper_id,
            doi=paper_data.get("doi") or None,
            arxiv_id=paper_data.get("arxiv_id") or None,
            title=title_lower,
            authors=paper_data.get("authors") or [],
            pub_date=pub_date_obj,
            venue=paper_data.get("venue") or None,
            abstract=paper_data.get("abstract") or None,
            source=paper_data.get("source"),
            retrieved_at=now,
            download_status="not_downloaded",
            ai_analysis_status="not_analyzed",
            pdf_path=paper_data.get("_pdf_path"),
            created_at=now,
            updated_at=now,
        )
        .on_conflict_do_nothing()
    )
    result = await db.execute(insert_stmt)
    if result.rowcount == 0:
        # 已存在，查找已有 ID
        existing = await db.execute(select(Paper.id).where(Paper.title == title_lower))
        row = existing.scalar_one_or_none()
        if row:
            paper_id = row

    # 建立课题-论文关联
    rel_stmt = (
        pg_insert(ProjectPaperRelation)
        .values(
            id=new_id("ppr"),
            project_id=project_id,
            paper_id=paper_id,
            is_valid=False,
            push_status=False,
            created_at=now,
            updated_at=now,
        )
        .on_conflict_do_nothing()
    )
    await db.execute(rel_stmt)
    await db.commit()

    return ok(AddPaperResponse(
        paper_id=paper_id,
        title=paper_data.get("title", "（解析中）"),
        status="added",
        message="论文已入库",
    ))


@router.get("/{project_id}/papers/{paper_id}")
async def get_paper(
    project_id: str,
    paper_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """获取论文完整详情 —— FR-007"""
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")

    paper, relation = await project_service.get_paper_with_relation(db, project_id, paper_id)
    if not paper or not relation:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "论文或关联记录不存在")

    return ok(PaperDetail(
        paper_id=paper.id,
        title=paper.title,
        authors=paper.authors,
        pub_date=str(paper.pub_date) if paper.pub_date else None,
        venue=paper.venue,
        abstract=paper.abstract,
        total_score=relation.total_score,
        reference_score=relation.reference_score,
        technical_score=relation.technical_score,
        is_valid=relation.is_valid,
        push_status=relation.push_status,
        download_status=paper.download_status,
        ai_analysis=paper.ai_analysis,
        doi=paper.doi,
        arxiv_id=paper.arxiv_id,
        source=paper.source,
        pdf_path=paper.pdf_path,
        text_path=paper.text_path,
        download_error=paper.download_error,
        ai_analysis_status=paper.ai_analysis_status,
        scoring_reason=relation.scoring_reason,
        pushed_at=relation.pushed_at,
        retrieved_at=paper.retrieved_at,
        created_at=paper.created_at,
        updated_at=paper.updated_at,
    ))


@router.patch("/{project_id}/papers/{paper_id}/score")
async def update_paper_score(
    project_id: str,
    paper_id: str,
    body: UpdatePaperScoreRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """手动修改论文评分，total_score 自动计算，is_valid 按 ≥7 规则更新 —— FR-007"""
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")

    if not (0 <= body.reference_score <= 5 and 0 <= body.technical_score <= 5):
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "评分须在 0-5 范围内")

    relation = await project_service.update_paper_score(
        db, project_id, paper_id,
        body.reference_score, body.technical_score, body.scoring_reason,
    )
    if not relation:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "论文关联不存在")

    return ok(PaperScoreResponse(
        paper_id=paper_id,
        reference_score=relation.reference_score,
        technical_score=relation.technical_score,
        total_score=relation.total_score,
        is_valid=relation.is_valid,
        scoring_reason=relation.scoring_reason,
    ))


@router.delete("/{project_id}/papers")
async def clear_papers(
    project_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """清空课题所有论文关联（FR-007）"""
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")
    count = await project_service.clear_papers(db, project_id)
    return ok({"message": "cleared", "removed_count": count})


@router.post("/{project_id}/archive")
async def archive_project(
    project_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """归档课题，状态变为 archived（FR-005）"""
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")
    if project.status in ("running", "paused"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "有任务正在运行，请先取消后再归档")
    if project.status == "archived":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "课题已归档")
    project = await project_service.archive_project(db, project)
    return ok(_project_to_detail(project))


@router.delete("/{project_id}/papers/{paper_id}")
async def remove_paper(
    project_id: str,
    paper_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")
    removed = await project_service.remove_paper(db, project_id, paper_id)
    if not removed:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "论文关联不存在")
    return ok({"message": "removed"})


# ── FR-008 任务管理 ────────────────────────────────────────────────────────────

@router.get("/{project_id}/status")
async def get_project_status(
    project_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """返回当前项目状态和最新阶段信息（FR-008）"""
    from sqlalchemy import select
    from app.models.stage import StageRecord

    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")

    latest_stage = await db.execute(
        select(StageRecord)
        .where(StageRecord.project_id == project_id)
        .order_by(StageRecord.started_at.desc())
        .limit(1)
    )
    stage = latest_stage.scalar_one_or_none()
    return ok({
        "project_id": project.id,
        "status": project.status,
        "mode": project.mode,
        "current_stage": stage.stage if stage else None,
        "stage_status": stage.status if stage else None,
    })


@router.post("/{project_id}/cancel")
async def cancel_task(
    project_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """取消当前运行中的任务，状态归 idle（FR-008）"""
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")
    if project.status not in ("running", "paused"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "当前无可取消的任务")
    project.status = "idle"
    await db.commit()
    return ok({"message": "cancelled"})


# ── FR-009 数据导出 ────────────────────────────────────────────────────────────

@router.post("/{project_id}/export")
async def export_data(
    project_id: str,
    body: ExportRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """异步导出论文库（Excel 或 PDF ZIP）"""
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")
    if body.export_type not in ("excel", "pdf_zip"):
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "无效的导出类型")

    # ── 查询论文 ─────────────────────────────────────────────────────────────────
    stmt = (
        select(Paper, ProjectPaperRelation)
        .join(ProjectPaperRelation, Paper.id == ProjectPaperRelation.paper_id)
        .where(ProjectPaperRelation.project_id == project_id)
        .order_by(ProjectPaperRelation.total_score.desc().nullslast())
    )
    if body.is_valid is not None:
        stmt = stmt.where(ProjectPaperRelation.is_valid == body.is_valid)
    if body.push_status is not None:
        stmt = stmt.where(ProjectPaperRelation.push_status == body.push_status)

    rows = (await db.execute(stmt)).all()

    # ── Excel 导出 ───────────────────────────────────────────────────────────────
    if body.export_type == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Papers"
        headers = ["标题", "作者", "期刊/会议", "发表日期", "总分", "DOI", "arXiv ID", "摘要"]
        ws.append(headers)
        for paper, rel in rows:
            ws.append([
                paper.title,
                "；".join(paper.authors) if paper.authors else "",
                paper.venue or "",
                str(paper.pub_date) if paper.pub_date else "",
                rel.total_score,
                paper.doi or "",
                paper.arxiv_id or "",
                (paper.abstract or "")[:500],
            ])
        # 自动列宽
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"papers_{project_id}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # ── PDF ZIP 导出 ─────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for paper, _ in rows:
            if paper.pdf_path and os.path.isfile(paper.pdf_path):
                zf.write(paper.pdf_path, arcname=os.path.basename(paper.pdf_path))
    buf.seek(0)
    filename = f"pdfs_{project_id}.zip"
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── FR-010 定时任务配置 ────────────────────────────────────────────────────────

@router.get("/{project_id}/schedule")
async def get_schedule(
    project_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")
    return ok(ScheduleConfigResponse(
        auto_push=project.auto_push,
        push_interval=project.push_interval,
        next_push_at=project.next_push_at,
    ))


@router.patch("/{project_id}/schedule")
async def update_schedule(
    project_id: str,
    body: ScheduleConfigRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    project = await project_service.get_project(db, project_id, current_user.id)
    if not project:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "课题不存在")
    if body.auto_push and body.push_interval not in (1, 3, 7, 14, 30):
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, "push_interval 须为 1/3/7/14/30 天")

    project = await project_service.update_schedule(db, project, body.auto_push, body.push_interval)
    return ok(ScheduleConfigResponse(
        auto_push=project.auto_push,
        push_interval=project.push_interval,
        next_push_at=project.next_push_at,
    ))


# ── FR-011 推荐模块 ────────────────────────────────────────────────────────────

recommendations_router = APIRouter(prefix="/recommendations", tags=["推荐模块"])


@recommendations_router.get("")
async def list_recommendations(
    content_type: str | None = Query(None),
    keyword: str | None = Query(None),
    order_by: str = Query("created_at"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select
    from app.models.user import User as UserModel

    total, items = await project_service.list_recommendations(db, content_type, keyword, order_by, page, page_size)

    user_ids = list({r.user_id for r in items})
    users_result = await db.execute(select(UserModel).where(UserModel.id.in_(user_ids)))
    users_map = {u.id: u for u in users_result.scalars().all()}

    return ok(RecommendationListResponse(
        total=total, page=page, page_size=page_size,
        items=[
            RecommendationItem(
                recommendation_id=r.id,
                user_id=r.user_id,
                username=users_map[r.user_id].username if r.user_id in users_map else "",
                content_type=r.content_type,
                title=r.title,
                content=r.content,
                tags=r.tags,
                contact_info=r.contact_info,
                view_count=r.view_count,
                like_count=r.like_count,
                created_at=r.created_at,
            )
            for r in items
        ],
    ))


@recommendations_router.post("", status_code=status.HTTP_201_CREATED)
async def create_recommendation(
    body: CreateRecommendationRequest,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    rec = await project_service.create_recommendation(db, current_user.id, body.model_dump())
    return ok({"recommendation_id": rec.id})


@recommendations_router.delete("/{rec_id}")
async def delete_recommendation(
    rec_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """软删除自己发布的推荐内容（FR-011）"""
    deleted = await project_service.delete_recommendation(db, rec_id, current_user.id)
    if not deleted:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "推荐内容不存在或无权删除")
    return ok({"message": "deleted"})


@recommendations_router.post("/{rec_id}/like")
async def like_recommendation(
    rec_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    rec = await project_service.toggle_like(db, rec_id, +1)
    if rec is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "推荐内容不存在")
    return ok({"rec_id": rec_id, "like_count": rec.like_count})
